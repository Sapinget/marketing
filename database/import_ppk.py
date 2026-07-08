#!/usr/bin/env python3
"""Import PPK.xlsx sheets into the Laravel SQLite database. Clears existing data first."""

import warnings
warnings.filterwarnings('ignore')

import sqlite3
import json
import hashlib
from datetime import datetime, date

XLSX_PATH = "/Users/serverbot/Library/CloudStorage/GoogleDrive-official@purapuraponsel.com/Shared drives/PURA PURA PONSEL/DATA/DASHBOARD/MARKETING/app-script/PPK.xlsx"
DB_PATH   = "/Users/serverbot/Library/CloudStorage/GoogleDrive-official@purapuraponsel.com/Shared drives/PURA PURA PONSEL/DATA/DASHBOARD/MARKETING/database/database.sqlite"

import openpyxl
print("Loading workbook…")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

def make_source_id(prefix, raw_id, counter):
    if raw_id is not None and str(raw_id).strip() not in ('', 'None', 'nan'):
        return str(raw_id).strip()
    return f"{prefix}-import-{counter}"

def to_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s in ('None', 'nan'):
        return None
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def to_int(val):
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(',', '')))
    except (ValueError, TypeError):
        return 0

def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ('', 'None', 'nan') else s

def sheet_rows(sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping")
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        data.append(d)
    return headers, data

now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

con = sqlite3.connect(DB_PATH)
cur = con.cursor()

# ── CLEAR ALL DATA ────────────────────────────────────────────────────────────
print("\nClearing existing data…")
tables = [
    'master_plans', 'distributions', 'analytics', 'marketing_settings',
    'marketing_excel_rows', 'stock_names', 'unboxing', 'story_schedules', 'calendar_events',
    'ideation', 'program_promo', 'sell_out_targets', 'ads_performance',
    'harga_kompetitor', 'orderan_online', 'unit_ditanya', 'claim_garansi',
    'keep_barang', 'lpjk', 'lpjk_detail',
]
for t in tables:
    cur.execute(f"DELETE FROM {t}")
    cur.execute(f"DELETE FROM sqlite_sequence WHERE name='{t}'")
con.commit()
print(f"  Cleared {len(tables)} tables")

inserted_total = 0

# ── Master_Plan ───────────────────────────────────────────────────────────────
_, rows = sheet_rows('Master_Plan')
print(f"\nMaster_Plan: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('KONTEN', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO master_plans
        (source_id,title,format_konten,platforms,colab,editor,talent,script,caption,status,tanggal_rencana,distribution_meta,link_drive,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Judul')), to_str(r.get('Format_Konten')),
          to_str(r.get('Platforms')), to_str(r.get('Colab')), to_str(r.get('Editor')), to_str(r.get('Talent')),
          to_str(r.get('Skrip')), to_str(r.get('Caption')), to_str(r.get('Status')),
          to_date(r.get('Tanggal_Rencana')), to_str(r.get('Distribution_Meta')),
          to_str(r.get('Link_Drive')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Distribution ──────────────────────────────────────────────────────────────
_, rows = sheet_rows('Distribution')
print(f"\nDistribution: {len(rows)} rows")
ins = skp = 0
seen = set()
for i, r in enumerate(rows):
    master_id = to_str(r.get('Master_ID')) or ''
    platform  = to_str(r.get('Platform')) or ''
    if platform in ('', 'contentType'):
        skp += 1
        continue
    key = (master_id, platform)
    if key in seen:
        skp += 1; continue
    seen.add(key)
    cur.execute("""
        INSERT OR IGNORE INTO distributions
        (master_id,title,platform,tanggal_publish,link,type,raw_payload,converted_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (master_id, to_str(r.get('Judul')), platform,
          to_date(r.get('Tanggal_Publish')), to_str(r.get('Link')), None,
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}, skipped duplicates {skp}")
inserted_total += ins

# ── Analytics ─────────────────────────────────────────────────────────────────
_, rows = sheet_rows('Analytics')
print(f"\nAnalytics: {len(rows)} rows")
ins = skp = 0
seen = set()
for i, r in enumerate(rows):
    master_id = to_str(r.get('Master_ID')) or ''
    platform  = to_str(r.get('Platform')) or ''
    if platform in ('', 'contentType'):
        skp += 1
        continue
    key = (master_id, platform)
    if key in seen:
        skp += 1; continue
    seen.add(key)
    cur.execute("""
        INSERT OR IGNORE INTO analytics
        (master_id,title,platform,tanggal_publish,views,likes,comments,shares,raw_payload,converted_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (master_id, to_str(r.get('Judul')), platform,
          to_date(r.get('Tanggal_Publish')),
          to_int(r.get('Views')), to_int(r.get('Likes')),
          to_int(r.get('Comments')), to_int(r.get('Shares')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}, skipped duplicates {skp}")
inserted_total += ins

# ── Settings ──────────────────────────────────────────────────────────────────
headers, rows = sheet_rows('Settings')
print(f"\nSettings: {len(rows)} rows, {len(headers)} columns")
ins = 0
# Each column = one setting key; collect non-null unique values per column
skip_cols = {'key', 'values', ''}
col_values = {h: [] for h in headers if h not in skip_cols}
col_seen   = {h: set() for h in headers if h not in skip_cols}
for r in rows:
    for h in col_values:
        v = to_str(r.get(h))
        if v and v not in col_seen[h]:
            col_values[h].append(v)
            col_seen[h].add(v)
# Also handle explicit key/values columns if present
explicit_kv = {}
if 'key' in headers and 'values' in headers:
    for r in rows:
        k = to_str(r.get('key'))
        v = to_str(r.get('values'))
        if k and v:
            try:
                parsed = json.loads(v)
                if isinstance(parsed, list):
                    explicit_kv[k] = parsed
                else:
                    explicit_kv[k] = [v]
            except (json.JSONDecodeError, ValueError):
                explicit_kv.setdefault(k, [])
                if v not in explicit_kv[k]:
                    explicit_kv[k].append(v)
# Merge explicit into col_values
for k, vals in explicit_kv.items():
    if k and vals:
        col_values[k] = vals

for key, vals in col_values.items():
    if not vals:
        continue
    cur.execute("""
        INSERT OR REPLACE INTO marketing_settings (key, "values", imported_at, created_at, updated_at)
        VALUES (?,?,?,?,?)
    """, (key, json.dumps(vals, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins} setting keys")
inserted_total += ins

# ── Konfigurasi_Bonus (BONUS_CONFIG + BUDGET_CONFIG → marketing_settings) ────
ws_bonus = wb['Konfigurasi_Bonus']
bonus_rows = list(ws_bonus.iter_rows(min_row=2, values_only=True))
ins = 0
for row in bonus_rows:
    key = str(row[0]).strip() if row[0] is not None else ''
    val = str(row[1]).strip() if row[1] is not None else ''
    if not key or not val:
        continue
    try:
        json.loads(val)
    except (json.JSONDecodeError, ValueError):
        continue
    cur.execute("""
        INSERT OR REPLACE INTO marketing_settings (key, "values", imported_at, created_at, updated_at)
        VALUES (?,?,?,?,?)
    """, (key, val, now, now, now))
    ins += 1
con.commit()
print(f"\nKonfigurasi_Bonus: inserted {ins} config keys")
inserted_total += ins

# ── Nama_Stock (stock_names) ─────────────────────────────────────────────────
_, rows = sheet_rows('Nama_Stock')
print(f"\nNama_Stock: {len(rows)} rows")
ins = 0
seen = set()
for r in rows:
    kategori = ' '.join((to_str(r.get('KATEGORI')) or '').upper().split())
    brand = ' '.join((to_str(r.get('BRAND')) or '').upper().split())
    seri = ' '.join((to_str(r.get('SERI')) or '').upper().split())
    key = (kategori, brand, seri)
    if key in seen:
        continue
    seen.add(key)
    cur.execute("""
        INSERT OR REPLACE INTO stock_names (source_id, kategori, brand, seri, imported_at, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?)
    """, (to_str(r.get('ID')), kategori, brand, seri, now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Unboxing ──────────────────────────────────────────────────────────────────
_, rows = sheet_rows('Unboxing')
print(f"\nUnboxing: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('UBX', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO unboxing (source_id,nama,editor,status,upload_date,link,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Nama')), to_str(r.get('Editor')), to_str(r.get('Status')),
          to_date(r.get('Upload_Date')), to_str(r.get('Link')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Story_Schedule ────────────────────────────────────────────────────────────
_, rows = sheet_rows('Story_Schedule')
print(f"\nStory_Schedule: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('STR', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO story_schedules (source_id,tanggal,jam,story,catatan,link,is_genap,status,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_date(r.get('Tanggal')), to_str(r.get('Jam')), to_str(r.get('Story')),
          to_str(r.get('Catatan')), to_str(r.get('Link')), to_str(r.get('is_genap')),
          to_str(r.get('Status')), json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Calendar_Events ───────────────────────────────────────────────────────────
_, rows = sheet_rows('Calendar_Events')
print(f"\nCalendar_Events: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('CAL', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO calendar_events (source_id,nama_event,tanggal,warna,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Nama_Event')), to_date(r.get('Tanggal')), to_str(r.get('Warna')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Program_Promo ─────────────────────────────────────────────────────────────
_, rows = sheet_rows('Program_Promo')
print(f"\nProgram_Promo: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('PRO', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO program_promo (source_id,kategori,program,warna,harga,periode,rules,benefit,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Kategori')), to_str(r.get('Program')), to_str(r.get('Warna')),
          to_int(r.get('Harga')), to_str(r.get('Periode')), to_str(r.get('Rules')),
          to_str(r.get('Benefit')), json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── SellOut_Target ────────────────────────────────────────────────────────────
_, rows = sheet_rows('SellOut_Target')
print(f"\nSellOut_Target: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('SOT', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO sell_out_targets
        (source_id,vendor,kategori,brand,seri,nama_produk,target_unit,bonus_nominal,realisasi_unit,periode_start,periode_end,catatan,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Vendor')), to_str(r.get('Kategori')), to_str(r.get('Brand')),
          to_str(r.get('Seri')), to_str(r.get('Nama_Produk')),
          to_int(r.get('Target_Unit')), to_int(r.get('Bonus_Nominal')), to_int(r.get('Realisasi_Unit')),
          to_date(r.get('Periode_Start')), to_date(r.get('Periode_End')), to_str(r.get('Catatan')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Report_Ads → ads_performance ─────────────────────────────────────────────
_, rows = sheet_rows('Report_Ads')
print(f"\nReport_Ads (ads_performance): {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('ADS', r.get('ID'), i)
    # Extract platform from ID suffix (e.g. "uuid_Instagram" → "Instagram")
    raw_id = to_str(r.get('ID')) or ''
    platform = None
    if '_' in raw_id:
        parts = raw_id.rsplit('_', 1)
        if len(parts) == 2 and parts[1]:
            platform = parts[1]
    cur.execute("""
        INSERT OR IGNORE INTO ads_performance
        (source_id,nama,id_ads,tanggal,biaya,sisa_saldo,kategori,platform,jangkauan,suka,komentar,share,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Nama')), to_str(r.get('ID_Ads')), to_date(r.get('Tanggal')),
          to_int(r.get('Biaya')),
          to_int(r.get('Sisa_Saldo')) if r.get('Sisa_Saldo') is not None else None,
          to_str(r.get('Kategori')), platform,
          to_int(r.get('Jangkauan')), to_int(r.get('Suka')),
          to_int(r.get('Komentar')), to_int(r.get('Share')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Harga_Kompetitor ──────────────────────────────────────────────────────────
_, rows = sheet_rows('Harga_Kompetitor')
print(f"\nHarga_Kompetitor: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('HK', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO harga_kompetitor
        (source_id,nama_produk,harga_distributor_1,harga_distributor_2,harga_kompetitor,margin_profit,harga_rencana_jual,tanggal_cek,catatan,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Nama_Produk')),
          to_int(r.get('Harga_Distributor_1')), to_int(r.get('Harga_Distributor_2')),
          to_int(r.get('Harga_Kompetitor')), to_int(r.get('Margin_Profit')),
          to_int(r.get('Harga_Rencana_Jual')), to_date(r.get('Tanggal_Cek')),
          to_str(r.get('Catatan')), json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Orderan_Online ────────────────────────────────────────────────────────────
_, rows = sheet_rows('Orderan_Online')
print(f"\nOrderan_Online: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('OO', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO orderan_online
        (source_id,tanggal,ecommerce,handle,nama,type_unit,harga_online,nominal_cair,status,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_date(r.get('TANGGAL')), to_str(r.get('ECOMMERCE')),
          to_str(r.get('HANDLE')), to_str(r.get('NAMA')), to_str(r.get('TYPE UNIT')),
          to_int(r.get('HARGA ONLINE')),
          to_int(r.get('NOMINAL CAIR')) if r.get('NOMINAL CAIR') is not None else None,
          to_str(r.get('STATUS')), json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Unit_Ditanya ──────────────────────────────────────────────────────────────
_, rows = sheet_rows('Unit_Ditanya')
print(f"\nUnit_Ditanya: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('UD', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO unit_ditanya
        (source_id,tanggal,kategori,brand,seri,kondisi,tipe,ditanya,available,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_date(r.get('TANGGAL')), to_str(r.get('KATEGORI')),
          to_str(r.get('BRAND')), to_str(r.get('SERI')), to_str(r.get('KONDISI')),
          to_str(r.get('TIPE')), to_int(r.get('DITANYA')), to_str(r.get('AVAILABLE')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Claim_Garansi_Asuransi ────────────────────────────────────────────────────
_, rows = sheet_rows('Claim_Garansi_Asuransi')
print(f"\nClaim_Garansi_Asuransi: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('CG', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO claim_garansi
        (source_id,nama_customer,no_service,no_transaksi,tanggal_masuk,wa_customer,tipe,seri,model,status,lokasi_klaim,tanggal_estimasi,tanggal_diambil,garansi,kerusakan,keterangan,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('NAMA_CUSTOMER')), to_str(r.get('NO_SERVICE')),
          to_str(r.get('NO_TRANSAKSI')), to_date(r.get('TANGGAL_MASUK')),
          to_str(r.get('WA_CUSTOMER')), to_str(r.get('TIPE')), to_str(r.get('SERI')),
          to_str(r.get('MODEL')), to_str(r.get('STATUS')), to_str(r.get('LOKASI_KLAIM')),
          to_date(r.get('TANGGAL_ESTIMASI')), to_date(r.get('TANGGAL_DIAMBIL')),
          to_str(r.get('GARANSI')), to_str(r.get('KERUSAKAN')), to_str(r.get('KETERANGAN')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Keep_Barang ───────────────────────────────────────────────────────────────
_, rows = sheet_rows('Keep_Barang')
print(f"\nKeep_Barang: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('KB', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO keep_barang
        (source_id,tanggal_keep,nama,nomor_hp,type_hp,dp_uang_muka,harga_jual,rencana_pengambilan,handle_by,status,tanggal_expired,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_date(r.get('TANGGAL_KEEP')), to_str(r.get('NAMA')),
          to_str(r.get('NOMOR_HP')), to_str(r.get('TYPE_HP')),
          to_int(r.get('DP_UANG_MUKA')), to_int(r.get('HARGA_JUAL')),
          to_date(r.get('RENCANA_PENGAMBILAN')), to_str(r.get('HANDLE_BY')),
          to_str(r.get('STATUS')), to_date(r.get('TANGGAL_EXPIRED')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Lpjk ──────────────────────────────────────────────────────────────────────
_, rows = sheet_rows('Lpjk')
print(f"\nLpjk: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('LPJK', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO lpjk
        (source_id,nama_event,tanggal,budget_rencana,realisasi_biaya,status,keterangan,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, to_str(r.get('Nama_Event')), to_date(r.get('Tanggal')),
          to_int(r.get('Budget_Rencana')), to_int(r.get('Realisasi_Biaya')),
          to_str(r.get('Status')), to_str(r.get('Keterangan')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

# ── Lpjk_Detail ───────────────────────────────────────────────────────────────
_, rows = sheet_rows('Lpjk_Detail')
print(f"\nLpjk_Detail: {len(rows)} rows")
ins = 0
for i, r in enumerate(rows):
    sid = make_source_id('LPJKD', r.get('ID'), i)
    cur.execute("""
        INSERT OR IGNORE INTO lpjk_detail
        (source_id,master_id,kategori,nama_pengeluaran,satuan,jumlah,total,bukti,raw_payload,imported_at,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (sid, str(r.get('Master_ID') or ''), to_str(r.get('Kategori')),
          to_str(r.get('Nama_Pengeluaran')), to_str(r.get('Satuan')),
          to_int(r.get('Jumlah')) or 1, to_int(r.get('Total')), to_str(r.get('Bukti')),
          json.dumps(r, default=str, ensure_ascii=False), now, now, now))
    ins += 1
con.commit()
print(f"  inserted {ins}")
inserted_total += ins

con.close()
print(f"\nDone. Total inserted: {inserted_total}")
